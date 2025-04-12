# %% Import libraries

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import is_date_format
from pandas import DataFrame
from pyhwpx import Hwp
from pyjosa.josa import Josa

# %% Open HWP application

print("Hello from autohwp!")

hwp = Hwp(visible=False)
print("Hwp version:", hwp.Version)
"""
[WinError 2] 지정된 파일을 찾을 수 없습니다 RegisterModule 액션을 실행할 수 없음. 개발자에게 문의해주세요.
#### 위 에러는 무시해도 실행 가능함 (이유 불명)

Hwp version: [11, 0, 0, 6402]
#### 한글 2020 버전
"""

# %% Open Template file

template_path = Path("template/contract.hwp").absolute()
print("Template path:", template_path)

hwp.open(str(template_path))
print("Opened document title:", hwp.get_title())
"""
Temp path: c:\\Users\\user\\workspace\\autohwp\\temp\\contract.hwp

빈 문서 1 - 한글
#### visible=False로 실행하면 백그라운드에서 빈 문서가 열림
"""

# %% Read field names from the document
# hwp.get_field_list()는 문서의 모든 필드(누름틀) 이름을 가져오는 메서드입니다.
# 필드이름은 "FieldName{{0}}" 형식으로 되어 있으며, {{0}}는 필드 인덱스입니다.
# 필드 인덱스는 문서에서 필드가 여러 개 있을 때, 각 필드를 구분하기 위한 숫자입니다.
# 만약 인덱스를 지정하지 않고 채워넣으면, 모든 동일한 이름의 필드에 같은 값이 채워집니다.

fields = hwp.get_field_list()
print("Total number of fields:", len(fields))
print("Field names in the document:")
for field in fields:
    print(field)

"""
책임교수명(본문){{0}}
계약당사자명(본문){{0}}
프로젝트명{{0}}
총 사업기간{{0}}
당해연도 사업기간{{0}}
계약시작일{{0}}
계약종료일{{0}}
총 계약금액{{0}}
월 계약금액{{0}}
급여일{{0}}
계약일{{0}}
산학협력단장명(서명란){{0}}
책임교수명(서명란){{0}}
주소{{0}}
계약당사자명(서명란){{0}}
휴대폰번호{{0}}
"""

# %% Read field values from Excel

"""
엑셀에서 사용되는 필드는 다음과 같습니다.
| 책임교수명 | 계약당사자 | 프로젝트명 | 총 사업기간 | 당해연도 사업기간 | 계약시작일 | 계약종료일 | 총 계약금액 | 월 계약금액 | 급여일 | 계약일 | 주소 | 휴대폰번호 |

* 산학협력단장명은 UI상에서 설정 가능하게 만들며, 고정값으로 데이터에 추가됨
"""

excel_path = Path("data/contract_fill.xlsx").absolute()
print("Excel path:", excel_path)

wb = load_workbook(excel_path, data_only=True, read_only=True)
ws = wb.active
if ws is None:
    raise ValueError("No active worksheet found in the Excel file.")
print(ws.title)  # 시트 이름 출력

# TODO: 추후 UI 설정에서 변경할 값들
column_row = 1  # 열 이름이 있는 행
start_row = 2  # 데이터 시작 행
primary_column = "성명"  # 기본 키 열 이름


# 열 이름이 있는 행에서 기본 키 열의 인덱스 찾기
header = [cell.value for cell in next(ws.iter_rows(min_row=column_row, max_row=column_row))]
primary_column_index = header.index(primary_column)


# %% Convert the column values to matching field format
# 엑셀에서 읽은 날짜 형식과 Python의 strftime 형식 매핑

# TODO: 추후 UI에서 엑셀날짜포맷을 읽어 사용자가 직접 날짜 포맷을 설정할 수 있도록 변경
excel_to_strftime = {
    r'yyyy"년"\ m"월"\ d"일";@': "%Y년 %#m월 %#d일",  # Unix/Linux/macOS: %Y년 %-m월 %-d일
    r"yyyy/mm/dd/": "%Y.%m.%d.",
    # 필요 시 추가 매핑
}


def format_cell(cell: "Cell | MergedCell") -> str | None:
    """셀의 서식을 확인해 날짜라면 포맷된 문자열 반환"""

    if cell.value is None:
        return None

    if is_date_format(cell.number_format) and isinstance(cell.value, datetime):
        fmt = excel_to_strftime.get(cell.number_format, "%Y-%m-%d")
        if fmt:
            return cell.value.strftime(fmt)

    return str(cell.value)


# Debug: Print Cell formats of first data row
# print("Cell formats of first data row:")
# for row in ws.iter_rows(min_row=start_row, max_row=start_row):
#     for cell in row:
#         if isinstance(cell, MergedCell):
#             print(f"Cell {cell.coordinate} is a merged cell.")
#         else:
#             print(f"Cell {cell.coordinate} - Value: {cell.value}, Format: {cell.number_format}")

# Cell formats of first data row:
# Cell A2 - Value: 김철수, Format: General
# Cell B2 - Value: B2015, Format: General
# Cell C2 - Value: 인공지능기술사업화연구소, Format: General
# Cell D2 - Value: 연구교수(선임급), Format: General
# Cell E2 - Value: 이상훈, Format: General
# Cell F2 - Value: AI 기반 로봇 제어기술연구, Format: General
# Cell G2 - Value: 2025-01-01 00:00:00, Format: yyyy/mm/dd/
# Cell H2 - Value: 2025-12-31 00:00:00, Format: yyyy/mm/dd/
# Cell I2 - Value: 2025-03-01 00:00:00, Format: yyyy/mm/dd/
# Cell J2 - Value: 2025-05-31 00:00:00, Format: yyyy/mm/dd/
# Cell K2 - Value: 2025-04-01 00:00:00, Format: yyyy"년"\ m"월"\ d"일";@
# Cell L2 - Value: 2025-05-30 00:00:00, Format: yyyy"년"\ m"월"\ d"일";@
# Cell M2 - Value: 2, Format: 0_);[Red]\(0\)
# Cell N2 - Value: 3200000, Format: _("₩"* #,##0_);_("₩"* \(#,##0\);_("₩"* "-"_);_(@_)
# Cell O2 - Value: 1352582, Format: _("₩"* #,##0_);_("₩"* \(#,##0\);_("₩"* "-"_);_(@_)
# Cell P2 - Value: 2705164, Format: _("₩"* #,##0_);_("₩"* \(#,##0\);_("₩"* "-"_);_(@_)
# Cell Q2 - Value: 25, Format: General
# Cell R2 - Value: 2025-01-01 00:00:00, Format: yyyy"년"\ m"월"\ d"일";@
# Cell S2 - Value: 서울특별시 강남구, Format: General
# Cell T2 - Value: 010-1234-5678, Format: General


# 빈 행도 None으로 모두 읽으므로, 기본 키에 해당하는 값이 None이 아닌 행만 읽도록 필터링 설정
# 데이터를 읽으면서, 셀의 서식을 확인하고 날짜라면 포맷된 문자열로 변환
fill_data = []
for row in ws.iter_rows(min_row=start_row):
    if row[primary_column_index].value is not None:
        fill_data.append([format_cell(cell) for cell in row])

print(f"Number of rows read: {len(fill_data)}")
# print("First row of data:")
# print(fill_data[0])

# %% Convert Excel data to DataFrame

# 데이터프레임 변환
df = DataFrame(fill_data, columns=header)

print(df.columns)
# print(df.dtypes)
# print(df.head())

# %% Convert the column values to matching field format

# TODO: 추후 UI에서 설정한 값으로 변경
CONST_SANHAK = "산학협력단장명"
df[CONST_SANHAK] = "김응태"  # 고정값으로 사용

# 본문 이름 조사 추가
# 한국공학대학교 산학협력단장과 연구책임교수 ○○○을 “학교”라 하고 (프로젝트)연구교수 ○○○을 “교원”이라 하여 다음과 같이 계약한다.
df["성명(본문)"] = df["성명"].apply(lambda x: Josa.get_full_string(x, "을"))
df["책임교수명(본문)"] = df["책임교수명"].apply(lambda x: Josa.get_full_string(x, "을"))

# 컬럼 묶기: "기간시작 ~ 기간종료"
df["총 사업기간"] = df["총 사업기간 시작"] + " ~ " + df["총 사업기간 종료"]
df["당해연도 사업기간"] = df["당해연도 사업기간 시작"] + " ~ " + df["당해연도 사업기간 종료"]

# 금액: 쉼표 구분 추가
# 숫자(정수)로 변환 후 포맷팅
df["총 계약금액"] = df["총 계약금액"].astype(int).apply(lambda x: f"{x:,}")
df["월 계약금액"] = df["월 계약금액"].astype(int).apply(lambda x: f"{x:,}")

print(df)
"""
    성명     사번                 소속         직급 책임교수명            프로젝트명  \
0  김철수  B2015       인공지능기술사업화연구소  연구교수(선임급)   이상훈  AI 기반 로봇 제어기술연구
1  이영희  B2074         ICT지능화융합센터  연구교수(책임급)   김지원    스마트팩토리 운영설계사업
2  최민호  C5813  GRAND-ICT 인력양성사업단      박사연구원   박정수         로봇센터구축사업

     총 사업기간 시작    총 사업기간 종료 당해연도 사업기간 시작 당해연도 사업기간 종료  ...          계약종료일  \
0  2025.01.01.  2025.12.31.  2025.03.01.  2025.05.31.  ...   2025년 5월 30일
1  2025.01.01.  2025.12.31.  2025.05.01.  2025.07.31.  ...   2025년 7월 31일
2  2025.01.01.  2025.12.31.  2025.06.01.  2025.09.30.  ...  2025년 12월 31일

  참여개월수  예산 편성금액   월 계약금액   총 계약금액 급여일           계약일          주소  \
0     2  3200000  1352582  2705164  25   2025년 1월 1일   서울특별시 강남구
1     2  6500000  2923323  5846646  25   2025년 5월 1일     경기도 성남시
2     4  3800000   842488  3369952  25  2025년 7월 31일  부산광역시 해운대구

           휴대폰번호 산학협력단장명
0  010-1234-5678     김응태
1  010-2345-6789     김응태
2  010-3456-7890     김응태

[3 rows x 21 columns]
"""

# %% Match the field names with the column names in the Excel file

# TODO: 추후 UI에서 필드 이름과 엑셀 파일의 열 이름을 매핑할 수 있도록 변경
# 필드 이름과 엑셀 파일의 열 이름을 매핑하는 딕셔너리
field_mapping = {  # 한글 필드이름: 엑셀 열 이름
    "책임교수명(본문){{0}}": "책임교수명(본문)",
    "계약당사자명(본문){{0}}": "성명(본문)",
    "프로젝트명{{0}}": "프로젝트명",
    "총 사업기간{{0}}": "총 사업기간",
    "당해연도 사업기간{{0}}": "당해연도 사업기간",
    "계약시작일{{0}}": "계약시작일",
    "계약종료일{{0}}": "계약종료일",
    "총 계약금액{{0}}": "총 계약금액",
    "월 계약금액{{0}}": "월 계약금액",
    "급여일{{0}}": "급여일",
    "계약일{{0}}": "계약일",
    "산학협력단장명(서명란){{0}}": "산학협력단장명",
    "책임교수명(서명란){{0}}": "책임교수명",
    "주소{{0}}": "주소",
    "계약당사자명(서명란){{0}}": "성명",
    "휴대폰번호{{0}}": "휴대폰번호",
}


# %% Fill in the fields with values

# TODO: 추후 UI에서 설정한 값으로 변경
CONST_FILENAME = "프로젝트연구교수 채용계약서"

save_folder = Path(f"{CONST_FILENAME}" + "_" + datetime.now().strftime("%Y%m%d")).absolute()
save_folder.mkdir(parents=False, exist_ok=True)  # 폴더가 없으면 생성
print("Save folder:", save_folder)

for index, row in df.iterrows():
    # pandas Hashable to int index
    idx = int(str(index)) + 1
    print(f"문서 만드는중... ({idx}/{len(df)})")

    # 각 행의 값을 필드에 채워넣음
    for field_name, column_name in field_mapping.items():
        # 필드 이름에 해당하는 엑셀 열 이름을 찾아서 값을 가져옴
        column_value = getattr(row, column_name)

        # 필드에 값을 채워넣음
        hwp.put_field_text(field_name, column_value)

    # 채워진 양식 문서를 각 행의 성명으로 저장
    save_path = save_folder / f"{CONST_FILENAME}_{row['성명']}.hwp"

    # HWP 문서 및 PDF로 저장
    hwp.save_as(str(save_path))
    print(f"작성문서 저장 완료: {save_path.name}")

    hwp.save_pdf_as_image(str(save_path.with_suffix(".pdf")), img_format="jpg")
    print(f"PDF 저장 완료: {save_path.with_suffix('.pdf').name}")


# %% Quit HWP application
hwp.quit()
