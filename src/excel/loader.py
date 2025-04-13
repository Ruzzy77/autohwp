import warnings
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame

from excel.formatter import format_cell


def load_worksheet(path: str, **kwargs) -> Worksheet:
    """
    엑셀 파일을 로드하고 워크시트를 반환합니다.

    Args:
        path (str): 엑셀 파일 경로
        **kwargs: 추가 설정 값

    Returns:
        Worksheet: 엑셀 워크시트 객체
    """

    wb = load_workbook(path, data_only=True, read_only=True)

    if "sheet_name" in kwargs:
        ws = wb[kwargs["sheet_name"]]
    else:
        ws = wb.active or wb.worksheets[0]

    if ws is None:
        raise ValueError("No active worksheet found in the Excel file.")

    return ws


def data_loader(
    worksheet: Worksheet,
    header_row: int = 1,
    start_row: int | None = None,
    end_row: int | None = None,
    key_columns: list[Any] | None = None,
) -> DataFrame:
    """
    엑셀 워크시트에서 데이터를 로드하여 데이터프레임을 반환합니다.

    Args:
        worksheet (Worksheet): 엑셀 워크시트 객체
        header_row (int): 헤더가 있는 행 번호
            - 행 번호는 1부터 시작합니다.
        start_row (int | None): 데이터 시작 행 번호
            - 지정되지 않으면 header_row + 1로 설정됩니다.
        end_row (int | None): 데이터 종료 행 번호
            - 지정되지 않으면 모든 행이 포함됩니다.
        key_columns (list[Any]): 기본키 열 이름의 리스트
            - 키 열이 정의되지 않으면 첫 번째 열이 기본키로 사용됩니다.

    Returns:
        DataFrame: 엑셀 데이터로 생성된 데이터프레임
    """

    # Suppress openpyxl warnings
    # UserWarning: Data Validation extension is not supported and will be removed
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    # 헤더 행 로드
    header = [
        cell.value for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row))
    ]

    # 기본키 열 인덱스 설정
    if key_columns is None:
        key_columns = [header[0]]
    key_column_indices = [header.index(col) for col in key_columns]

    # print(f"key_columns: {key_columns}")
    # print(f"key_column_indices: {key_column_indices}")

    # 데이터 시작 행 설정
    if start_row is None:
        start_row = header_row + 1

    if end_row is None:
        end_row = worksheet.max_row

    # 데이터 로드
    fill_data = []
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row):
        # 모든 기본키 열의 값이 None이 아닌 경우에만 추가
        if all(row[col].value is not None for col in key_column_indices):
            fill_data.append([format_cell(cell) for cell in row])

    if not fill_data:
        raise ValueError("No valid data found in the specified range.")

    return DataFrame(fill_data, columns=header)
